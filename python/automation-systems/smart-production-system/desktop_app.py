# make shortcut in terminal with this pyinstaller --icon="C:/Users/ab/Desktop/Raspberry_PI/Programming/Final_Code/LLGS_working_version/images/Lear_L_Logo-removebg-preview-removebg-preview.ico" "C:/Users/ab/Desktop/Raspberry_PI/Programming/Final_Code/LLGS_working_version/Excel_Loading_interface_next_step.py"
#!/usr/bin/python

import os
from datetime import datetime
import tkinter as tk
from tkinter import *
from tkinter import filedialog, messagebox, ttk
import signal
import openpyxl
import customtkinter as ctk
from openpyxl.styles import Border, Side, PatternFill, Alignment,Font
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
import paramiko
from PIL import Image
import json
# from openpyxl.worksheet.table import Table, TableStyleInfo

# Get the desktop directory
desktop_dir = os.path.join(os.path.expanduser("~"), "Desktop")

# Specify the file path for saving the workbook on the desktop
desktop_excel_path = os.path.join(desktop_dir, "LLGS_server_configuration.xlsx")
# N02, N01, L01, L02, U01, U02
identifier={"R8B2-14K024-ZLX": ('IP RH L663','I3R'), "R8B2-14401-ZLX": ('IP RH L663', 'I3R'), 
            "S8E2-14401-ZAD": ('IP RH L663','I3R'), "S8E2-14401-YAD": ('IP RH L663', 'I3R'),
            "SY42-14401-ZC": ('IP RH L663','I3R'), "RY42-14K024-YA": ('IP RH L663', 'I3R')}
################################## Sending txt file to raspberry pi ###############################################
class send_ssh:
    def __init__(self, localpath):
        self.localpath = localpath
        self.load_ipaddress_password_dict=self.load_ipaddress_password_dict()

    def load_ipaddress_password_dict(self):
        try:
            with open("ipaddress_password_dict.json", "r") as file:
                self.ipaddress_password_dict = json.load(file)
        except FileNotFoundError:
            self.ipaddress_password_dict = {}
        
    def send_file_to_raspberry_pi(self, combobox_structure_shortcut):
        pi_ip_address = self.ipaddress_password_dict[combobox_structure_shortcut][0]
        password = self.ipaddress_password_dict[combobox_structure_shortcut][1]
        local_path = self.localpath
        remotepath=f"/home/pi/Documents/LLGS"
        try:
            # Create SSH client
            ssh_client = paramiko.SSHClient()
            ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())

            # Connect to the Raspberry Pi via SSH
            ssh_client.connect(hostname=pi_ip_address, username='pi', password=password)

            print("Connected to Raspberry Pi via SSH")

            # Create SFTP client
            sftp_client = ssh_client.open_sftp()

            # Upload the local file to the Raspberry Pi
            sftp_client.put(local_path, remotepath)

            print("File uploaded successfully to Raspberry Pi in structure: ",combobox_structure_shortcut,
                  "\nwith Ip address: ", pi_ip_address,
                  "\nStored at: ", remotepath)
            
            # Close the SFTP session
            sftp_client.close()
            # Close the SSH session
            ssh_client.close()
            messagebox.showinfo("SSH Upload",f"File uploaded successfully to raspberry pi with IP: {pi_ip_address}.\nFile Stored at: {remotepath}")
            

        except Exception as e:
            print("Error:", str(e))
            messagebox.showerror("SSH Conncetion Failed",f"Failed to establish connection, with error: {e}")

    def printStructure(self, structure_name_shortcut):
        print(self.ipaddress_password_dict[structure_name_shortcut][0])
        print(self.ipaddress_password_dict[structure_name_shortcut][1])
        print(self.localpath)

#############################################################################################################
################################## Reading Excel and creating the txt file ##################################
class create_config_txt: 
    #define the init function which will run automatically
    def __init__(self, excel_path, txt_file_path, desktop_excel_path):
        self.excel_path = excel_path
        self.txt_file_path=txt_file_path
        self.desktop_excel_path= desktop_excel_path
        #check if file has 40 wires in Wire sheet
        workbook = openpyxl.load_workbook(self.excel_path)
        ws = workbook["Wires"]
       
      #sort unique wires by length and append them to txt file
    
    def sort_by_fourth_column(self):
        os.system("taskkill /f /im excel.exe")
        print("Excel terminated using taskkill")
        # Path to your Excel file
        file_path = self.desktop_excel_path

        workbook = openpyxl.load_workbook(self.excel_path)
        worksheet = workbook["Wires"]

        # Initialize a list to store data from the worksheet
        data = []
        
        # Iterate through the rows starting from the third row (excluding the header)
        for row in worksheet.iter_rows(min_row=3, values_only=True):
            data.append(row)  # Append entire row as a list to the data list
        # Sort the data by the values in the fourth column (index 3)
        # Sort the data by the values in the V column (index 22)
        sorted_data = sorted(data, key=lambda x: x[3], reverse=True)
        print(sorted_data)
        # Extract the second column (index 1) from the sorted data
        sorted_wires = [row[1] for row in sorted_data]
        
        unique_wires = "unique_wires:"
        for key in sorted_wires:
            unique_wires+= '"' + key+'",'
            
        unique_wires=unique_wires[:-1]+"wires_code:"
        
        #append the unique wires into the txt file
        #write splices unique names into the txt file
        with open(self.txt_file_path, "a") as f:
            f.write(unique_wires)
        return sorted_wires

    #this function creates the binary 1, 0 list for the wires in or out from the excel
    def generate_binary_list(self):
        wb = openpyxl.load_workbook(self.desktop_excel_path)
        # Select the sheet
        ws = wb["Splices"]
        
        unique_wires= self.sort_by_fourth_column()
        # Get the number of rows and columns in the range
        nb_row = ws.max_row

        for col in range(1, ws.max_column + 1):
            if ws.cell(row=2, column=col).value == ws.cell(row=3, column=2).value:
        # Return the column number of the match - 1
                nb_column = col - 1
    
        # Create the binary list
        binary_list = []
        
        # Iterate over each row in the range
        for row in ws.iter_rows(min_row=3, max_row=nb_row+1, min_col=11, max_col=nb_column, values_only=True):
            if row and any(row):  # Check if the row is not empty and contains at least one non-empty value
                # Initialize a binary row with 'x' for each wire not present, and '1' for each wire present
                binary_row = ['1' if wire in row else 'x'  for wire in unique_wires]
            
                # Append the binary row to the binary list
                binary_list.append(binary_row)
        
        #get unique splices 
        unique_splices=[]
        for row in ws.iter_rows(min_row=3, max_row=nb_row,min_col=2, values_only=True):
            if row and any(row):  # Check if the row is not empty and contains at least one non-empty value
            
                # Append the binary row to the binary list
                unique_splices.append(row[0])
                
        binary_dic= dict(zip(unique_splices,binary_list))
     
        return binary_dic
    
    def read_excel_server_code(self):
        splices_data = {}
        
        # Load the Excel workbook
        workbook = openpyxl.load_workbook(self.excel_path)
        sheet = workbook["Splices"]
        
        # Iterate through the rows and extract the data as a tuple and then transform it into a list and then a dictionary
        for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row-1, values_only=True):
            key = row[1]  # Assuming the key is in the 2nd column of each row
            quantity=row[3]
            family=row[0]
            
            # Store the data in the dictionary
            splices_data[key] = {"quantity": quantity , "family" : family}
            
        
        return splices_data 
    #read data to create server message, and append unique splices to txt fileat top
    def excel_data_for_server(self):
        
        ############################### adding the G2Dtest splice line ###############################
        wb = openpyxl.load_workbook(self.excel_path)
        # Select the sheet
        sheet = wb["Splices"]

        quantity_family= self.read_excel_server_code()
        number_splices= len(quantity_family)
        #print("number_splices",number_splices)
        
        unique_wires= self.sort_by_fourth_column()
 
        family = sheet.cell(row=3, column=1).value
        
        #index of the one to be written inot the sheet for every testing splice each time increase it by one
        
        testing_splice_line= [family, "G2Dtest", "", 1, "", 0, 3.3, 3.3,20, 20]
        
        # Extend my_list with the elements of another_list
        testing_splice_line.extend(unique_wires)

        none_zero_list= (number_splices+1)*[""]+[1]
    
        testing_splice_line.extend(none_zero_list)
        sheet.append(testing_splice_line)
        
        sheet.cell(row=2, column=sheet.max_column).value= "G2Dtest"
        wb["Wires"].cell(row=2, column=wb["Wires"].max_column+1).value="G2Dtest"
        i=0
        for wire in unique_wires:
            i+=1
            wb["Wires"].cell(row=2+i, column=wb["Wires"].max_column).value=1
        # Save the workbook
        wb.save(self.desktop_excel_path)
        
        ########################################### add the 40 lines for testing ################################
        #errors if there are not enough values declared
        testing_splices_rows = []
    
        shift=1
        for wire in unique_wires:
            # Change the first letter of each string to 'T'
            test = 'T' + wire[1:]
            
            testing_splices_lines = [family, test, "", 1, "", 0, 3.3, 3.3, 1, 0, wire]
            
            testing_splices_lines.extend((39+number_splices+1+shift)*[""]+[1])
            # Append testing_splices_lines to testing_splices_rows
            testing_splices_rows.append(testing_splices_lines)
            
            # Update the cell in row 2 and column sheet.max_column with the value of test
            sheet.cell(row=2, column=sheet.max_column+1).value = test
            wb["Wires"].cell(row=2, column=wb["Wires"].max_column+1).value = test
            i=0
            for key in unique_wires:
                i+=1
                if wb["Wires"].cell(row=2+i, column=2).value== "G"+test[1:]:
                    
                    wb["Wires"].cell(row=2+i, column=wb["Wires"].max_column).value=1
                    #print("G"+test[1:], wb["Wires"].cell(row=2+i, column=2).value )
            shift+=1
            
        for row in testing_splices_rows:
            sheet.append(row)
        ############################################ Wire Sheet formating ################################
        
        #Borders for all cells except header
        # Define the range of rows and columns to set solid borders
        start_row = 3  # Start row
        end_row =  wb["Wires"].max_row+1  # End row
        start_col = 1  # Start column
        end_col =   wb["Wires"].max_column+1  # End column

        # Define the border style (solid)
        border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))

        # Iterate through the specified range of cells and apply border style
        for row in range(start_row, end_row):
            for col in range(start_col, end_col):
                cell =  wb["Wires"].cell(row=row, column=col)
                cell.border = border

            #color the headers with blue and change the rotation of the text   
    # Define the fill color (blue) for headers
        blue_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')

        # Apply blue fill color to cells in row 2 and set rotation
        for col in range(1, wb["Wires"].max_column + 1):
            cell = wb["Wires"].cell(row=2, column=col)
            cell.fill = blue_fill
            cell.alignment = Alignment(textRotation=90)
            cell.font = Font(color="FFFFFF")  # Set font color to white
            
        # Define the fill color for conditional formatting
        empty_color = '808080'
        one_color = 'CCFFCC'

        ####### Define the range of rows to apply conditional formatting
        start_row = 3  # Start row
        end_row = wb["Wires"].max_row + 1   # End row
        start_column = 26  # Start column
        end_column = wb["Wires"].max_column + 1  # End column (for example, column Z)

        # Add conditional formatting rule for cells with value 1
        one_rule = CellIsRule(operator='equal', formula=['1'], fill=PatternFill(start_color=one_color, end_color=one_color, fill_type='solid'))
        wb["Wires"].conditional_formatting.add(f'{get_column_letter(start_column)}{start_row}:{get_column_letter(end_column)}{end_row}', one_rule)

        ####### fill the blank cells with gray
        start_row = 3  # Start row
        end_row =  wb["Wires"].max_row + 1  # End row
        start_column = 26  # Start column
        end_column = wb["Wires"].max_column + 1  # End column (for example, column J)

        # Iterate through the specified range of cells
        for row in range(start_row, end_row ):
            for col in range(start_column, end_column ):
                cell = wb["Wires"].cell(row=row, column=col)
                if not cell.value:  # Check if cell is empty
                    cell.fill = PatternFill(start_color=empty_color, end_color=empty_color, fill_type='solid')

        # Save the workbook
        wb.save(self.desktop_excel_path)
        
        ################################################## Formating ##################################################
        # Load the Excel file
        wb = openpyxl.load_workbook(self.desktop_excel_path)
        ws = wb["Splices"]

    #Borders for all cells except header
        # Define the range of rows and columns to set solid borders
        start_row = 3  # Start row
        end_row = ws.max_row+1  # End row
        start_col = 1  # Start column
        end_col =  ws.max_column+1  # End column

        # Define the border style (solid)
        border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))

        # Iterate through the specified range of cells and apply border style
        for row in range(start_row, end_row):
            for col in range(start_col, end_col):
                cell = ws.cell(row=row, column=col)
                cell.border = border

    #fill the first column rows with yellow
        # Define the range of rows to fill with yellow color
        start_row = 3  # Start row
        end_row = ws.max_row+1  # End row

        # Define the fill color (yellow)
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        # Iterate through the specified range of rows and apply yellow fill to the cells in the first column
        for row in range(start_row, end_row):
            cell = ws.cell(row=row, column=1)  # First column
            cell.fill = yellow_fill
                 
    #color the headers with blue and change the rotation of the text   
    # Define the fill color (blue) for headers
        blue_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')

        # Apply blue fill color to cells in row 2 and set rotation
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=2, column=col)
            cell.fill = blue_fill
            cell.alignment = Alignment(textRotation=90)
            cell.font = Font(color="FFFFFF")  # Set font color to white
            
        # Define the fill color for conditional formatting
        empty_color = '808080'
        one_color = 'CCFFCC'

        ####### Define the range of rows to apply conditional formatting
        start_row = 3  # Start row
        end_row = ws.max_row + 1   # End row
        start_column = 51  # Start column
        end_column = ws.max_column + 1  # End column (for example, column Z)

        # Add conditional formatting rule for cells with value 1
        one_rule = CellIsRule(operator='equal', formula=['1'], fill=PatternFill(start_color=one_color, end_color=one_color, fill_type='solid'))
        ws.conditional_formatting.add(f'{get_column_letter(start_column)}{start_row}:{get_column_letter(end_column)}{end_row}', one_rule)

        ####### fill the blank cells with gray
        start_row = 3  # Start row
        end_row =  ws.max_row + 1  # End row
        start_column = 51  # Start column
        end_column = ws.max_column + 1  # End column (for example, column J)

        # Iterate through the specified range of cells
        for row in range(start_row, end_row ):
            for col in range(start_column, end_column ):
                cell = ws.cell(row=row, column=col)
                if not cell.value:  # Check if cell is empty
                    cell.fill = PatternFill(start_color=empty_color, end_color=empty_color, fill_type='solid')
    
        # Save the workbook
        wb.save(self.desktop_excel_path)
        
        #empty the txt file
        with open(self.txt_file_path, "w") as f:
                pass
            
        formatted_time = '{:02d}:{:02d}'.format(datetime.now().hour, datetime.now().minute)
        formatted_date = '{}/{}/{}'.format(datetime.now().month, datetime.now().day, datetime.now().year)

        formatted_datetime = formatted_time + " " + formatted_date
        unique_splices=f"{formatted_date}\nunique_splices:"
        
        splices_data={}
        
        wb = openpyxl.load_workbook(self.desktop_excel_path)
        # Select the sheet
        sheet = wb["Splices"]
        #print("sheet max row: ", sheet.max_row)
        #print("sheet max column: ", sheet.max_column)
        
        # Iterate through the rows and extract the data as a tuple and then transform it into a list and then a dictionary
        for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row, values_only=True):
            key = row[1]  # Assuming the key is in the second column of each row | first column in excel is row[0]
            None_1_tuple = row[-(sheet.max_row-2):sheet.max_column+1]
            quantity=row[3]
            family=row[0]
            
            # Convert None_1_tuple to a list with 'x' instead of None in one line
            x_1_list = ['x' if item is None else item for item in None_1_tuple]
        
            # Store the data in the dictionary
            splices_data[key] = {"on_off": x_1_list, "quantity": quantity , "family" : family}
            
            #the list for all unique splices
            unique_splices+='"'+key+'",'
            
        #write splices unique names into the txt file
        with open(self.txt_file_path, "a") as f:
            f.write(unique_splices[:-1])
            
        return splices_data
    
    #this function creates the module_list code and stores it in the txt file
    def server_msg_creation(self):
        #read on_off state, quantity, family
        splices_data = self.excel_data_for_server()
        
        #the path to the configuration file
        txt_file_path=self.txt_file_path
        
        #date to out i the the server qr
        Current_Date= str(datetime.now().year)+str(datetime.now().month).zfill(2)+ str(datetime.now().day).zfill(2)+ str(datetime.now().hour).zfill(2)+ str(datetime.now().minute).zfill(2)+ str(datetime.now().second).zfill(2)
        bsn_number=0
        
        for key in splices_data:
            bsn_number= int(bsn_number)+1
            
            x_1_data= splices_data[key]["on_off"]
            x_1_data_no_comma= ''.join(str(e) for e in x_1_data)
            
            #bsn counter format as 9 digits
            bsn_number = '{:09d}'.format(bsn_number)
        
            quantity= splices_data[key]["quantity"]
            #family and short name of family
            server_family=identifier[splices_data[key]["family"]][0]
            server_family_short= identifier[splices_data[key]["family"]][1]
         
            #"LJS@000000001@001@I3R@LOT@50@.@20240416221107@.@.@.@IP RH L663@0.0000@.@1xxxxxxxxxxxxxxxxxxxxx@000000001="
            #Concatenate all elements into the data
            server_msg=("LJS@"+str(bsn_number)+"@001@"+server_family_short+"@LOT@"+str(quantity)+"@.@"+str(Current_Date)+"@.@.@.@"+server_family+"@0.0000@.@"+str(x_1_data_no_comma)+"@000000001=")
            txt_configuration_msg_1= server_msg.replace('LJS','"LJS').replace('"LJS@000000001','module_list:"LJS@000000001').replace('x@000000001=','x@000000001=",').replace('1@000000001=','1@000000001="')
            
            #write concatenated data into the txt file
            with open(txt_file_path, "a") as f:
                f.write(txt_configuration_msg_1)
                
        #print("module_list_code is written into txt file at " + txt_file_path)

    #this function reads the excel data related to the module_list code creation process
    def read_excel_wires_code(self):
        splices_data = {}
        
        # Load the Excel workbook
        workbook = openpyxl.load_workbook(self.desktop_excel_path)
        sheet = workbook["Splices"]
        
        # Iterate through the rows and extract the data as a tuple and then transform it into a list and then a dictionary
        for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row, values_only=True):
            key = row[1]  # Assuming the key is in the 2nd column of each row
            quantity=row[3]
            family=row[0]
            
            # Store the data in the dictionary
            splices_data[key] = {"quantity": quantity , "family" : family}
            
        
        return splices_data 
    
    #this function creates the wires code and stores it in the txt file
    def wires_code(self):
        #LJS@000000001@001@I3R@LOT@50@.@20240416220616@.@.@.@IP RH L663@0.0000@G2D215@111xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx@000000001
        
        #the path to the configuration file
        txt_file_path=self.txt_file_path
        
        #binary dictionnary that contains 1 for exists and x for none existing, key is splice name
        x_1_dict= self.generate_binary_list()
        
        splices_info_dict = self.read_excel_wires_code()
     
        #date to out i the the server qr
        Current_Date= str(datetime.now().year)+str(datetime.now().month).zfill(2)+ str(datetime.now().day).zfill(2)+ str(datetime.now().hour).zfill(2)+ str(datetime.now().minute).zfill(2)+ str(datetime.now().second).zfill(2)
        
        bsn_number=0
        server_msg=""
        for key in splices_info_dict:
            
            x_1_data_no_comma= ''.join(str(e) for e in x_1_dict[key])
            
            bsn_number= int(bsn_number)+1
            
            #bsn counter format as 9 digits
            bsn_number = '{:09d}'.format(bsn_number)
        
            quantity= splices_info_dict[key]["quantity"]
            #family and short name of family
            excel_family=  splices_info_dict[key]["family"]
        
            server_family=identifier[excel_family][0]
            server_family_short= identifier[excel_family][1]
        
            #LJS@000000001@001@I3R@LOT@50@.@20240416220616@.@.@.@IP RH L663@0.0000@G2D215@111xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx@000000001
            #Concatenate all elements into the data
            server_msg+=("LJS@"+str(bsn_number)+"@001@"+server_family_short+"@LOT@"+str(quantity)+"@.@"+str(Current_Date)+"@.@.@.@"+server_family+"@0.0000@"+key+"@"+str(x_1_data_no_comma)+'@000000001=",')
        txt_configuration_msg_2= server_msg.replace('LJS','"LJS')
            
        #print(txt_configuration_msg_2)
            #write concatenated data into the txt file
        with open(txt_file_path, "a") as f:
            f.write(txt_configuration_msg_2[:-1])
                
        print("data is written into txt file at " + txt_file_path)
        messagebox.showinfo("Empty Field",message="data is written into txt file at " + txt_file_path) 
        #open genrated txt file
        if os.path.exists(txt_file_path):
            os.startfile(txt_file_path)
        else:
            print("No Data is sent to the configuration file.")
        #open generated excel file
        if os.path.exists(desktop_excel_path):
            os.startfile(desktop_excel_path)
        else:
            print("No Data is sent to the configuration file.")

#############################################################################################################
#close the interface function
def kill_script():
    os.kill(os.getpid(), signal.SIGTERM)

#window customisation function
def window_header(root, page_title):
    # Getting screen width and height of display
    width = root.winfo_screenwidth()
    height = root.winfo_screenheight()

    # Setting tkinter window size
    root.geometry("%dx%d" % (width, height))
    
    # Maximize the window
    root.after(0, lambda: root.state("zoomed"))
    
    # Top section
    top_section = ctk.CTkFrame(root, height=100, border_width=0.5)
    top_section.pack(side=tk.TOP, fill=tk.X)
    
    # Titles
    label_function(page_title, 0.5, 0.5,top_section, 32)
    label_function("LLGS", 0.95, 0.5, top_section, 32)
    '''
    # Open and resize the dark image
    dark_image_path = "images/Lear_logo_light.png"
    dark_image = Image.open(dark_image_path)
    dark_image = dark_image.resize((145, 45))
    
    # Open and resize the light image
    light_image_path = "images/lear.png"
    light_image = Image.open(light_image_path)
    light_image = light_image.resize((145, 45))
 
    # Lear image
    img = ctk.CTkImage(dark_image=dark_image, light_image=light_image, size=(145, 45))
    my_label = ctk.CTkLabel(top_section, text="", image=img)
    my_label.pack(side="left", padx=3, pady=3)
    '''
#asseign the excel path written manually to the excel_file_path variable
def assign_excel_path(ent_box):
    path_text = ent_box.get()
    excel_path.set(path_text)  # excel_path to the entered value 

#asseign the path written manually to the text_file_path variable
def assign_txt_path(ent_box):
    path_text = ent_box.get()
    txt_file_path.set(path_text) # txt_file_path to the entered value

#button customisation function
def button_function(instance, name, action, relx, rely, WIDTH=200, height=45, border_width=2, size=18, fg_color="#cc0000"):
    button = ctk.CTkButton(master=instance, text=name,width=WIDTH, corner_radius=5,border_width=border_width,height=height, command=action, font=("Helvetica", size, "bold"), fg_color=fg_color, hover_color="#F10F0F", border_color="#b00101")
    button.place(relx=relx, rely=rely, anchor=CENTER)

def label_function(name, relx, rely, root, size=24, bgcolor="transparent"):
# excel path label
    scan_text = ctk.CTkLabel(root, text=name, font=("Helvetica", size,'bold'), fg_color=bgcolor, padx=0, corner_radius=0)
    scan_text.place(relx=relx, rely=rely, anchor="center")
    return scan_text

def entry_function(text_variable, relx, rely, placeholder):
    excel_path_box = ctk.CTkEntry(master=root,height=2, border_width=2,border_color="#b8142a",placeholder_text=placeholder, corner_radius=5, textvariable=text_variable, width=300, font=("Arial", 18), justify="center")
    excel_path_box.place(relx=relx, rely=rely, anchor="center")
    return excel_path_box

def get_excel_path(excel_lebel_update):
    # Open file dialog to select a file
    global excel_path
    excel_path= StringVar()
    
    excel_path = filedialog.askopenfilename()
    
    if excel_path:
        
        #write the path browsed into the entry box
        update_entry_text(excel_lebel_update,excel_path)
    
    else:
        messagebox.showerror("Empty Field",message="enter valid path") 

#this is the action for browsing for the target txt file
def get_txt_file_path(txt_label_update):
    # Open file dialog to select a file
    global txt_file_path
    txt_file_path= StringVar()
    
    txt_file_path = filedialog.askopenfilename()
    if txt_file_path:
        
      
        
        #write the path browsed into the entry box
        update_entry_text(txt_label_update, txt_file_path)
        
    else:
        messagebox.showerror("Empty Field",message="select the target txt file") 

#this is the action of the create button
def start_creation(excel_path, txt_file_path, combobox_structure_shortcut): 

    print("start creation excel path : ", excel_path)
    print("start creation txt pat : ", txt_file_path)
    
    if excel_path and txt_file_path and combobox_structure_shortcut!="...":
        
        #first class: instance of create_config_txt class with excel path, and txt path
        config_class= create_config_txt(excel_path, txt_file_path, desktop_excel_path)
        
        #create server message and append it to the txt file, also appends unique splices
        config_class.server_msg_creation()
        #create wires code and append it to the txt file
        config_class.wires_code()
    
        #second class: instance of send_ssh class with
        send_class= send_ssh(txt_file_path)
        send_class.printStructure(combobox_structure_shortcut)
        #send file to raspberry pi via ssh connection
        send_class.send_file_to_raspberry_pi(combobox_structure_shortcut)
        
    elif not excel_path or not txt_file_path or combobox_structure_shortcut=="...":
        messagebox.showerror("Empty Field",message="Make sure to fill all necessary fields.")

#upadates the enrtry box after browsing for the file
def update_entry_text(label_box,excel_path):
    paste_text= excel_path
    # Update the label with the new text
    label_box.configure(text=paste_text)

#################################################################################################################################
################################################# Configuration Password winodw #################################################
#the password page verification
class password_page:
    def __init__(self):
        clear_window(root)

        self.window = root
        self.window.title("Structure Credentials Manager")
        window_header(self.window, "Password Verification")

        frame = ctk.CTkFrame(master=self.window, border_width=3) 
        frame.pack(pady=260,padx=260,fill='both',expand=True) 

        label_function("Enter Your Password:", 0.5, 0.2, frame)

        user_pass= ctk.CTkEntry(master=frame, placeholder_text="Password",show="*",  border_width=2,border_color="#b8142a", corner_radius=5, width=300, font=("Arial", 18), justify="center") 
        user_pass.place(relx=0.5, rely=0.35, anchor="center") 

        user_pass.focus_force()
        
        button_function(frame, "Login", lambda: self.login(user_pass), 0.5, 0.5, 400)

        button_function(frame, "Back", lambda: path_loading_page(), 0.5, 0.65, 400,45,2, 18, "transparent")
     
        root.bind('<Return>', lambda event: self.login(user_pass))
    
    def login(self, user_pass): 
        password = "llgsPi7"
        if user_pass.get() == password:
            PasswordManagerApp() 
        else : 
            messagebox.showwarning(title='Wrong password',message='Please check your password') 
            clear_window(root)
            password_page()

############################################################################################################################
################################################# Cnfiguration Edit window #################################################
#the edit window used to edit the content of the structure credentials
class PasswordManagerApp:
    def __init__(self):
        clear_window(root)
        
        self.parent = root
        self.parent.title("Structure Credentials Manager")
        window_header(root, "Workstation Configuration (Raspberry Pi devices)")
        self.load_data()
        
        frame = ctk.CTkFrame(master=self.parent, border_width=3) 
        frame.pack(pady=160,padx=160,fill='both',expand=True) 
        
        label_function("Add New Structure", 0.5, 0.2, frame)
        self.add_button = button_function(frame, "Add", self.add_entry_window, 0.5, 0.3, 500)
        
        label_function("Delete Existing Structure", 0.5, 0.4, frame)
        self.delete_button = button_function(frame, "Delete", self.delete_entry_window, 0.5, 0.5, 500)
        
        label_function("Exit configuration Page", 0.5, 0.6, frame)
        self.exit_button = button_function(frame, "Back", self.destroyparent, 0.5, 0.7, 500,45,2, 18, "transparent")

    def destroyparent(self):
        path_loading_page()
    
    def load_data(self):
        try:
            with open("ipaddress_password_dict.json", "r") as file:
                self.passwords = json.load(file)
        except FileNotFoundError:
            self.passwords = {}

    def save_data(self):
        with open("ipaddress_password_dict.json", "w") as file:
            json.dump(self.passwords, file)

    def add_entry_window(self):
        clear_window(root)
        self.load_data()
        self.add_window = self.parent
        self.add_window.title("Add Entry")
        window_header(root, "Workstation Configuration (Raspberry Pi devices)")
        
        self.new_key = tk.StringVar()
        self.new_ip = tk.StringVar()
        self.new_password = tk.StringVar()

        frame = ctk.CTkFrame(master=self.parent, border_width=3) 
        frame.pack(pady=160,padx=160,fill='both',expand=True) 
       
        label_function("New Key:", 0.5, 0.1 , frame)
        box1 = ctk.CTkEntry(master=frame,textvariable=self.new_key, border_width=1,border_color="#b8142a",placeholder_text="new structure name", corner_radius=30, width=300, font=("Arial", 18), justify="center")
        box1.place(relx=0.5, rely=0.2, anchor="center")
        
        label_function("New IP Address:", 0.5, 0.3 , frame)
        box2 = ctk.CTkEntry(master=frame, textvariable=self.new_ip, border_width=1,border_color="#b8142a",placeholder_text="new IP Address", corner_radius=30, width=300, font=("Arial", 18), justify="center")
        box2.place(relx=0.5, rely=0.4, anchor="center")
       
        label_function("New Password:", 0.5, 0.5 , frame)
        box3 = ctk.CTkEntry(master=frame, textvariable=self.new_password, border_width=1,border_color="#b8142a",placeholder_text="New Password", corner_radius=30, width=300, font=("Arial", 18), justify="center")
        box3.place(relx=0.5, rely=0.6, anchor="center")

        button_function(frame, "Add", self.add_entry, 0.5, 0.75,400)
        button_function(frame, "Back", lambda: self.__init__(), 0.5, 0.9, 400,45,2, 18, "transparent")

    def add_entry(self):
        key = self.new_key.get()
        ip = self.new_ip.get()
        password = self.new_password.get()

        if key in self.passwords:
            messagebox.showerror("Error", "Key already exists.")
        else:
            self.passwords[key] = (ip, password)
            self.save_data()
            messagebox.showinfo("Success", "Entry added successfully.")
            clear_window(root)
            self.add_entry_window()

    def delete_entry_window(self):  
        clear_window(root)
        self.load_data()
        
        self.delete_window = self.parent
        window_header(root, "Workstation Configuration (Raspberry Pi devices)")
        self.delete_window.title("Delete Entry")

        self.key_to_delete = tk.StringVar(value=list(self.passwords.keys()))
        self.selected_key_to_delete = tk.StringVar()

        frame = ctk.CTkFrame(master=self.parent, border_width=3) 
        frame.pack(pady=260,padx=260,fill='both',expand=True)
        
        label_function("Select Key:", 0.5, 0.2 , frame)
        #tk.OptionMenu(self.delete_window, self.selected_key_to_delete, *list(self.passwords.keys())).pack()
        combobox=ctk.CTkOptionMenu(master=frame, values=list(self.passwords.keys()),
                                   variable=self.selected_key_to_delete, button_color="#6A6B6F",
                                   fg_color="#cc0000", font=("Helvetica", 18, "bold"), button_hover_color="black",
                                   dropdown_font=("Helvetica", 18, "bold"), anchor="center", width=200, height=45
                                   
        )
        
        combobox.place(relx=0.5, rely=0.4, anchor="center")
        combobox.set("...")
        
        button_function(frame, "Delete", self.delete_entry, 0.5, 0.6,300)
       
        button_function(frame, "Back", lambda: self.__init__(), 0.5, 0.8,300,45,2, 18,  "transparent")
    
    def delete_entry(self):
        selected_key = self.selected_key_to_delete.get()
        del self.passwords[selected_key]
        self.save_data()
        msg= messagebox.showinfo("Success", "Entry deleted successfully.")
        clear_window(root)
        self.delete_entry_window()

def clear_window(window):
    # Iterate through all the widgets in the window and destroy them
    for widget in window.winfo_children():
        widget.destroy()

#function to shorten the button twi lines in the configuration windows
def button_function_pad(instance, name, command, pady):
    button = ctk.CTkButton(instance, text=name, command=command,font=("Arial", 18) , fg_color="#b8142a")
    button.pack(pady=pady)

#function that centers the windows to the screen
def CenterWindowToDisplay(Screen: ctk, width: int, height: int, scale_factor: float = 1.0):
    """Centers the window to the main display/monitor"""
    screen_width = Screen.winfo_screenwidth()
    screen_height = Screen.winfo_screenheight()
    x = int(((screen_width/2) - (width/2)) * scale_factor)
    y = int(((screen_height/2) - (height/1.5)) * scale_factor)
    return f"{width}x{height}+{x}+{y}"

#########################################################################################################################
################################################# Main Interface Window #################################################
#the first window that appears after runing the application
def path_loading_page():
    clear_window(root)
    # Create a frame for the window header
    window_header(root, "Create Configuration File")
    
    ipaddress_password_dict={}
    ipaddress_password_dict= read_or_create_json_file("ipaddress_password_dict.json")

                                           #labels, entry boxes combboxes
                                           
    frame = ctk.CTkFrame(master=root, border_width=3) 
    frame.pack(pady=20,padx=20,fill='both',expand=True) 
    
    # excel path label
    label_function("Excel File", 0.4, 0.2 , frame)

    # excel path entry
    excel_path_update=label_function("", 0.5, 0.3 , frame, 16, "#1B1B1B")

    # txt path label
    label_function("Data txt file (empty one)", 0.4, 0.4 , frame)
    
    # txt path entry 
    txt_file_path_update= label_function("", 0.5, 0.5 , frame, 16, "#1B1B1B")
    
     # modify confugurations
    label_function("Edit Workstation's Configurations", 0.4, 0.6, frame)
    
    # Chose Structure label for combbox
    label_function("Workstation", 0.4, 0.75, frame)
    # combbox
    combobox = ctk.CTkComboBox(master=frame, values=list(ipaddress_password_dict.keys()),
                               justify="center", button_color="#6A6B6F", border_color='#6A6B6F',
                                   fg_color="#cc0000", font=("Helvetica", 18, "bold"), button_hover_color="black",
                                   dropdown_font=("Helvetica", 18, "bold"), width=200, height=45)
    
    combobox.place(relx=0.7,rely=0.75, anchor="center")
   
    combobox.set("...")
                                          #buttons
                                
    # excel file browsing button
    button_function(frame, "Add Excel source file", lambda: get_excel_path(excel_path_update), 0.7, 0.2)

    # txt browsing button
    button_function(frame, "Add txt target file", lambda: get_txt_file_path(txt_file_path_update), 0.7, 0.4)

    #configuration 
    button_function(frame, "Configurations", lambda: password_page(), 0.7, 0.6)
    
    #second method of enter the path through button
    button_function(frame, "Create and Load Configuration File", lambda: start_creation(excel_path, txt_file_path, combobox.get()), 0.5, 0.9, 600, 50)
    
    #Exit button
    button_function(frame, "X", lambda: kill_script(), 0.98, 0.03,20,10,0,25, "transparent")

def read_or_create_json_file(file_path):
    """
    Read data from a JSON file or create one with an empty dictionary if not found.

    Args:
    - file_path (str): The path to the JSON file.

    Returns:
    - dict: The data read from the JSON file as a dictionary.
    """
    if os.path.exists(file_path):
        try:
            with open(file_path, "r") as file:
                data = json.load(file)
            return data
        except json.JSONDecodeError:
            print(f"Error: Failed to decode JSON from file '{file_path}'.")
            return {}
    else:
        print(f"File '{file_path}' not found. Creating a new file with an empty dictionary.")
        with open(file_path, "w") as file:
            json.dump({}, file)
        return {}

# Create the main tk instance
root = ctk.CTk()
ctk.set_appearance_mode("dark")

# Set the window title
root.title("Configuration Application")
root.geometry(CenterWindowToDisplay(root, 450, 250, root._get_window_scaling()))
root.wm_iconbitmap("C:/Users/ab/Desktop/Raspberry_PI/Programming/Final_Code/LLGS_working_version/images/lear.ico")

#call the first page window function
path_loading_page()

# Run the Tkinter event loop
root.mainloop()